from typing import Optional
import os
import datetime
import io
import requests
import pandas as pd
import numpy as np
import yfinance as yf
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import numbers

# ===== USER CONFIGURATION =====================================
# 1️⃣  Telegram credentials
TELEGRAM_TOKEN = "8097099702:AAEBs7N8UkpFmOgRxjS60-EdOnoRSBYVUMw"   # <-- keep private
CHAT_ID        = "1736831659"                                       # <— your user/chat id

# 2️⃣  Core filter thresholds
MIN_MARKET_CAP   = 300_000_000      # > $300 M
MIN_PRICE        = 1.0              # >= $1 so we ignore penny stocks
MIN_ADR_PCT      = 3.0              # > 3 % 14‑day Average Daily Range
# ==============================================================

TODAY     = datetime.datetime.now().strftime("%Y-%m-%d")
FILENAME  = f"filtered_stocks_{TODAY}.xlsx"

# -----------------------------------------------------------------------------
# TELEGRAM HELPERS
# -----------------------------------------------------------------------------
SEND_MSG_URL  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
SEND_DOC_URL  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"

def tg_send_message(text:str):
    try:
        resp = requests.post(SEND_MSG_URL, data={"chat_id": CHAT_ID, "text": text})
        print("Telegram msg:", resp.json())
    except Exception as e:
        print("‼️  Telegram message failed:", e)

def tg_send_file(path: str, caption: str = ""):
    try:
        with open(path, "rb") as fp:
            resp = requests.post(
                SEND_DOC_URL,
                data={"chat_id": CHAT_ID, "caption": caption},
                files={"document": (os.path.basename(path), fp)}
            )
        print("Telegram file:", resp.json())
    except Exception as e:
        print("‼️  Telegram file failed:", e)

# -----------------------------------------------------------------------------
# DATA COLLECTION
# -----------------------------------------------------------------------------
MARKET_TICKERS = ["SPY", "^VIX", "^TNX"]

def get_market_indicators():
    data = yf.download(MARKET_TICKERS,
                       period="2d",
                       interval="1d",
                       group_by="ticker",
                       auto_adjust=True)
    changes = {}
    for tic in MARKET_TICKERS:
        try:
            last = data[tic]["Close"].iloc[-1]
            prev = data[tic]["Close"].iloc[-2]
            key  = tic.strip("^")
            changes[key] = round(((last - prev) / prev) * 100, 2)
        except Exception:
            # we still want the key even if lookup fails
            changes[tic.strip("^")] = None
    return changes


    return changes

# universe = S&P 500 + a few liquid extras
EXTRA_TICKERS = ["TSLA", "NVDA", "AMD", "SHOP", "PLTR", "BABA", "F", "GM"]

def load_stock_universe():
    sp500 = pd.read_html("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies")[0]["Symbol"].tolist()
    tickers = list(set(sp500 + EXTRA_TICKERS))

    rows = []
    lookup_failed = []
    for tk in tickers:
        try:
            st = yf.Ticker(tk)
            info = st.fast_info  # lightweight
            hist = st.history(period="15d")
            if len(hist) < 14:
                continue

            price      = info.get("last_price", np.nan)
            prev_close = info.get("previous_close", np.nan)
            change_pct = round(((price - prev_close) / prev_close) * 100, 2)

            adr = (hist["High"] - hist["Low"]).rolling(14).mean().iloc[-1]
            adr_pct = round((adr / price) * 100, 2) if price else 0

            rows.append({
                "Ticker": tk,
                "Price": round(price, 2),
                "Prev Close": round(prev_close, 2),
                "% Change": change_pct,
                "ADR%": adr_pct,
                "Market Cap": info.get("market_cap", np.nan),
                "Sector": info.get("sector", "N/A"),
                "Industry": info.get("industry", "N/A"),
            })
        except Exception:
            lookup_failed.append(tk)
    if lookup_failed:
        print("Skipped tickers (data issues):", ",".join(lookup_failed))
    return pd.DataFrame(rows)

# -----------------------------------------------------------------------------
# FILTERING LOGIC
# -----------------------------------------------------------------------------

def apply_filters(df: pd.DataFrame, spy_move: Optional[float]):
    # Basic thresholds
    df = df[(df["Market Cap"] > MIN_MARKET_CAP) &
            (df["Price"] > MIN_PRICE) &
            (df["ADR%"] > MIN_ADR_PCT)]

    # SPY-relative day change
    if spy_move is not None:
        if spy_move > 0:
            df = df[df["% Change"] > spy_move]
        elif spy_move < 0:
            df = df[df["% Change"] < spy_move]

    # Sort strongest first (by absolute % Change)
    df = df.sort_values("% Change", ascending=False, ignore_index=True)
    return df

# -----------------------------------------------------------------------------
# EXCEL OUTPUT
# -----------------------------------------------------------------------------

def format_market_cap(val):
    try:
        if val >= 1_000_000_000_000:
            return f"{val/1e12:.2f}T"
        if val >= 1_000_000_000:
            return f"{val/1e9:.2f}B"
        if val >= 1_000_000:
            return f"{val/1e6:.2f}M"
    except Exception:
        return val
    return val


def save_excel(df: pd.DataFrame, indicators: dict):
    wb = Workbook()
    # Sheet 1 – market summary
    ws1 = wb.active
    ws1.title = "Market Summary"
    ws1.append(list(indicators.keys()))
    ws1.append(list(indicators.values()))

    # Sheet 2 – filtered stocks
    ws2 = wb.create_sheet("Filtered Stocks")
    df_fmt = df.copy()
    df_fmt["Market Cap"] = df_fmt["Market Cap"].apply(format_market_cap)

    for r in dataframe_to_rows(df_fmt, index=False, header=True):
        ws2.append(r)

    # % formatting
    for hdr in ("% Change", "ADR%"):
        col = [c.column_letter for c in ws2[1] if c.value == hdr][0]
        for row in range(2, ws2.max_row + 1):
            ws2[f"{col}{row}"].number_format = numbers.BUILTIN_FORMATS[10]  # 0.00%

    wb.save(FILENAME)

# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    market_moves = get_market_indicators()
    print("Market moves:", market_moves)

    stocks         = load_stock_universe()
    filtered       = apply_filters(stocks, market_moves.get("GSPC"))

    save_excel(filtered, market_moves)

    # Build Telegram summary
    msg = (
        f"📊 SPY: {market_moves.get('GSPC')}%  | VIX: {market_moves.get('VIX')}%  | TNX: {market_moves.get('TNX')}%\n"
        f"✅ {len(filtered)} stocks matched all filters.\n"
        "🏆 Top 5 by % Change:\n" +
        "\n".join(
            f"{row['Ticker']}: {row['% Change']:.2f}%" for _, row in filtered.head(5).iterrows()
        )
    )
    tg_send_message(msg)

    # Attach the Excel file
    tg_send_file(FILENAME, caption=FILENAME)

    print("✅ All done!")

